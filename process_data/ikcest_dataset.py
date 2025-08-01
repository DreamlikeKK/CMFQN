import copy
import cv2
import torch
import torch.utils.data as data
import data_util as data_util
import torchvision.transforms.functional as F
import albumentations as A
from PIL import Image
import os
import openpyxl
import numpy as np
from tqdm import tqdm

torch.cuda.set_device(1)

class ikcest_dataset(data.Dataset):

    def __init__(self,
                 dataset='ikcest_dataset', image_size=224, 
                 is_train=True,
                 is_val=False,
                 is_test=False,
                 l='cn'
                 ):
        super(ikcest_dataset, self).__init__()
        self.l = l
        self.not_valid_set = set()
        self.label_ambiguity = []
        self.is_train = is_train
        self.is_val = is_val
        self.is_test = is_test
        root_path = '/data/data'
        self.root_path = root_path

        self.root_path_ambiguity = root_path
        self.index = 0
        self.label_dict = []
        self.image_size = image_size
        self.transform_just_resize = A.Compose(
            [
                A.Resize(always_apply=True, height=image_size, width=image_size)
            ]
        )
        self.transform = A.Compose(
            [
                A.HorizontalFlip(p=0.5),
                A.OneOf(
                    [
                        A.CLAHE(always_apply=False, p=0.25),
                        A.RandomBrightnessContrast(always_apply=False, p=0.25),
                        A.Equalize(always_apply=False, p=0.25),
                        A.RGBShift(always_apply=False, p=0.25),
                    ]
                ),
                A.OneOf(
                    [
                        A.ImageCompression(always_apply=False, quality_lower=60, quality_upper=100, p=0.2),
                        A.GaussianBlur(always_apply=False, p=0.2),
                        A.GaussNoise(always_apply=False, p=0.2),
                        A.ISONoise(always_apply=False, p=0.2)
                    ]
                ),
                A.Resize(always_apply=True,height=image_size, width=image_size)
            ]
        )
        from torchvision import transforms
        self.clip_transform = transforms.Compose([
            transforms.Resize((224, 224)), 
            transforms.ToTensor(),
            transforms.Normalize(
                mean=[0.48145466, 0.4578275, 0.40821073],  
                std=[0.26862954, 0.26130258, 0.27577711]
            )
        ])
        
        relative_path = self.root_path.lstrip('/')
        full_path = os.path.join(os.getcwd(), relative_path)
        if self.is_train:
            wb = openpyxl.load_workbook(f"{full_path}/train_{self.l}_datasets_news_data.xlsx")
        elif self.is_val:
            wb = openpyxl.load_workbook(f"{full_path}/val_{self.l}_datasets_news_data.xlsx")
        elif self. is_test:
            wb = openpyxl.load_workbook(f"{full_path}/test_{self.l}_datasets_news_data.xlsx")

        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row 

        fake_news_num = 0
        for i in tqdm(range(2, rows + 1)):
            label = int(sheet['C' + str(i)].value)
            label = 1 if label == 0 else 0
            fake_news_num += label

        downsample_rate = fake_news_num / (rows - fake_news_num)
        print(f"Downsample rate: {downsample_rate}")

        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet['A' + str(i)].value)
            label = int(sheet['C' + str(i)].value)
            label = 1 if label==0 else 0
            content = str(sheet['B' + str(i)].value)
            record = {}
            record['images'] = images_name
            record['label'] = label
            record['content'] = content
            self.label_dict.append(record)

        assert len(self.label_dict)!=0, 'Error: GT path is empty.'

    def __getitem__(self, index):
        find_path, img_GT = False, None
        record = self.label_dict[index]
        images, label, content = record['images'], record['label'], record['content']
        GT_path = images
        relative_path = self.root_path.lstrip('/')
        
        if self.is_train:
            img_f_path = 'train/img'
        elif self.is_val:
            img_f_path = 'val/img'
        elif self.is_test:
            img_f_path = 'test/img'

        full_path = os.path.join(relative_path, img_f_path)

        GT_path = "{}/{}".format(full_path, GT_path)

        img_GT = cv2.imread(GT_path, cv2.IMREAD_COLOR)
                
        if img_GT.ndim == 2:
            img_GT = np.expand_dims(img_GT, axis=2)
        if img_GT.shape[2] > 3:
            img_GT = img_GT[:, :, :3]

        img_GT = data_util.channel_convert(img_GT.shape[2], 'RGB', [img_GT])[0]

        img_pil = Image.fromarray(cv2.cvtColor(img_GT, cv2.COLOR_BGR2RGB))
        img_clip = self.clip_transform(img_pil)  # [C, 224, 224],

        if not self.is_train:
            img_GT_augment = self.transform_just_resize(image=copy.deepcopy(img_GT))["image"]
        else:
            img_GT_augment = self.transform(image=copy.deepcopy(img_GT))["image"]

        img_GT = self.transform_just_resize(image=copy.deepcopy(img_GT))["image"]
        img_GT = img_GT.astype(np.float32) / 255.
        img_GT_augment = img_GT_augment.astype(np.float32) / 255.

        if img_GT.shape[2] == 3:
            img_GT = img_GT[:, :, [2, 1, 0]]
        if img_GT_augment.shape[2] == 3:
            img_GT_augment = img_GT_augment[:, :, [2, 1, 0]]

        img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()
        img_GT_augment = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT_augment, (2, 0, 1)))).float()

        return (content, img_GT, img_GT_augment, label, 0, index, img_clip), (GT_path)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t

